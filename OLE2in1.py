import os
import tkinter as tk
from tkinter import filedialog, simpledialog
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string

def get_user_input(task_name):
    root = tk.Tk()
    root.withdraw()

    process_range = simpledialog.askstring("Input", f"Enter process range for {task_name} (e.g., CV21:DM200):")
    if not process_range or ':' not in process_range:
        print(f"Invalid process range for {task_name}. Exiting.")
        return None, None

    criteria_range = simpledialog.askstring("Input", f"Enter criteria range for {task_name} (e.g., CV21:CV200):")
    if not criteria_range or ':' not in criteria_range:
        print(f"Invalid criteria range for {task_name}. Exiting.")
        return None, None

    return process_range, criteria_range

def parse_range(range_str):
    start, end = range_str.split(':')
    start_col, start_row = start[:2], int(start[2:])
    end_col, end_row = end[:2], int(end[2:])
    return start_col, start_row, end_col, end_row

def process_excel_files(target_folder, task_name, process_range, criteria_range):
    print(f"Starting the process for {task_name}...")
    print(f"Process range: {process_range}")
    print(f"Criteria range: {criteria_range}")

    proc_start_col, proc_start_row, proc_end_col, proc_end_row = parse_range(process_range)
    crit_start_col, crit_start_row, crit_end_col, crit_end_row = parse_range(criteria_range)

    merge_wb = Workbook()
    merge_ws = merge_wb.active
    merge_ws.title = f"Merged Data {task_name}"

    files_processed = 0
    sheets_processed = 0
    rows_merged = 0

    for filename in os.listdir(target_folder):
        if filename.endswith('.xlsx') and not filename.startswith('~$'):
            file_path = os.path.join(target_folder, filename)
            print(f"Processing file: {filename}")

            try:
                wb = load_workbook(file_path, data_only=True)

                for sheet_name in wb.sheetnames:
                    if sheet_name not in ["index", "list", "setting", "TEMPLATE", "STUDENTINFO"]:
                        print(f"  Processing sheet: {sheet_name}")
                        ws = wb[sheet_name]

                        last_row = crit_start_row - 1
                        for row in range(crit_start_row, crit_end_row + 1):
                            if ws[f'{crit_start_col}{row}'].value is not None and ws[f'{crit_start_col}{row}'].value != "":
                                last_row = row

                        if last_row >= crit_start_row:
                            print(f"    Copying range {proc_start_col}{proc_start_row}:{proc_end_col}{last_row}")

                            # Copy non-blank rows to merge sheet
                            for row in range(proc_start_row, last_row + 1):
                                if any(ws.cell(row=row, column=col).value not in (None, "")
                                       for col in range(column_index_from_string(proc_start_col), column_index_from_string(proc_end_col) + 1)):
                                    rows_merged += 1
                                    for col in range(column_index_from_string(proc_start_col), column_index_from_string(proc_end_col) + 1):
                                        source_value = ws.cell(row=row, column=col).value
                                        merge_ws.cell(row=rows_merged, column=col - column_index_from_string(proc_start_col) + 1, value=source_value)

                                    # Add sheet name in column S
                                    merge_ws.cell(row=rows_merged, column=19, value=sheet_name)

                            sheets_processed += 1

                wb.close()
                files_processed += 1

            except Exception as e:
                print(f"Error processing file {filename}: {str(e)}")

    # Concatenate values from columns P and S and replace values in column P
    for row in range(1, merge_ws.max_row + 1):
        p_value = merge_ws.cell(row=row, column=16).value
        s_value = merge_ws.cell(row=row, column=19).value
        if p_value and s_value:
            concat_value = f"{p_value}-{s_value}"
            merge_ws.cell(row=row, column=16, value=concat_value)

    merge_path = os.path.join(target_folder, f"merge_{task_name}.xlsx")
    try:
        merge_wb.save(merge_path)
        print(f"Merged data saved to {merge_path}")
    except Exception as e:
        print(f"Error saving merged file: {str(e)}")

    print(f"Process completed for {task_name}. Files processed: {files_processed}, Sheets processed: {sheets_processed}")
    print(f"Total non-blank rows merged: {rows_merged}")

if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()
    
    target_folder = filedialog.askdirectory(title="Select Target Folder")
    if not target_folder:
        print("No folder selected. Exiting.")
        exit()

    ole_process_range, ole_criteria_range = get_user_input("OLE")
    displine_process_range, displine_criteria_range = get_user_input("Displine")

    if ole_process_range and ole_criteria_range:
        process_excel_files(target_folder, "OLE", ole_process_range, ole_criteria_range)

    if displine_process_range and displine_criteria_range:
        process_excel_files(target_folder, "Displine", displine_process_range, displine_criteria_range)

    input("Press Enter to exit...")