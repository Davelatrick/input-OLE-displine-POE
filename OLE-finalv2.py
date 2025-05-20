import os
import tkinter as tk
from tkinter import ttk, filedialog, simpledialog, messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill
import pandas as pd

# Function to get user input for process and criteria ranges
def get_user_input(task_name):
    root = tk.Tk()
    root.withdraw()

    process_range = simpledialog.askstring("Input", f"Enter process range for {task_name} (e.g., CV21:DM200):")
    if not process_range or ':' not in process_range:
        print(f"Invalid process range for {task_name}. Exiting.")
        return None, None

    criteria_range = simpledialog.askstring("Input", f"Enter criteria range for {task_name} (e.g., CV21:CV200):")
    if not criteria_range or ':' not in criteria_range:
        print(f"Invalid criteria range for {task_name}. Exiting.")
        return None, None

    return process_range, criteria_range

# Function to parse range strings
def parse_range(range_str):
    start, end = range_str.split(':')
    start_col, start_row = start[:2], int(start[2:])
    end_col, end_row = end[:2], int(end[2:])
    return start_col, start_row, end_col, end_row

# Function to process Excel files and generate merged files
def process_excel_files(target_folder, task_name, process_range, criteria_range):
    print(f"Starting the process for {task_name}...")
    print(f"Process range: {process_range}")
    print(f"Criteria range: {criteria_range}")

    proc_start_col, proc_start_row, proc_end_col, proc_end_row = parse_range(process_range)
    crit_start_col, crit_start_row, crit_end_col, crit_end_row = parse_range(criteria_range)

    merge_wb = Workbook()
    merge_ws = merge_wb.active
    merge_ws.title = f"Merged Data {task_name}"

    files_processed = 0
    sheets_processed = 0
    rows_merged = 0

    for filename in os.listdir(target_folder):
        if filename.endswith('.xlsx') and not filename.startswith('~$'):
            file_path = os.path.join(target_folder, filename)
            print(f"Processing file: {filename}")

            try:
                wb = load_workbook(file_path, data_only=True)

                for sheet_name in wb.sheetnames:
                    if sheet_name not in ["index", "list", "setting", "TEMPLATE", "STUDENTINFO"]:
                        print(f"  Processing sheet: {sheet_name}")
                        ws = wb[sheet_name]

                        last_row = crit_start_row - 1
                        for row in range(crit_start_row, crit_end_row + 1):
                            if ws[f'{crit_start_col}{row}'].value is not None and ws[f'{crit_start_col}{row}'].value != "":
                                last_row = row

                        if last_row >= crit_start_row:
                            print(f"    Copying range {proc_start_col}{proc_start_row}:{proc_end_col}{last_row}")

                            # Copy non-blank rows to merge sheet
                            for row in range(proc_start_row, last_row + 1):
                                if any(ws.cell(row=row, column=col).value not in (None, "")
                                       for col in range(column_index_from_string(proc_start_col), column_index_from_string(proc_end_col) + 1)):
                                    rows_merged += 1
                                    for col in range(column_index_from_string(proc_start_col), column_index_from_string(proc_end_col) + 1):
                                        source_value = ws.cell(row=row, column=col).value
                                        merge_ws.cell(row=rows_merged, column=col - column_index_from_string(proc_start_col) + 1, value=source_value)

                                    # Add sheet name in column S
                                    merge_ws.cell(row=rows_merged, column=19, value=sheet_name)

                            sheets_processed += 1

                wb.close()
                files_processed += 1

            except Exception as e:
                print(f"Error processing file {filename}: {str(e)}")

    # Concatenate values from columns P and S and replace values in column P
    for row in range(1, merge_ws.max_row + 1):
        p_value = merge_ws.cell(row=row, column=16).value
        s_value = merge_ws.cell(row=row, column=19).value
        if p_value and s_value:
            concat_value = f"{p_value}-{s_value}"
            merge_ws.cell(row=row, column=16, value=concat_value)

    # Additional concatenation for "Displine" task: concatenate I and S into I
    if task_name == "Displine":
        for row in range(1, merge_ws.max_row + 1):
            i_value = merge_ws.cell(row=row, column=9).value  # Column I
            s_value = merge_ws.cell(row=row, column=19).value  # Column S
            if i_value and s_value:
                concat_value = f"{i_value}-{s_value}"
                merge_ws.cell(row=row, column=9, value=concat_value)
    # === ADD THE FOLLOWING CODE ===
    from openpyxl.styles import PatternFill

    # 1. Insert header row at the top with 't1' across all columns with data
    merge_ws.insert_rows(1)
    max_col = merge_ws.max_column
    for col in range(1, max_col + 1):
        cell = merge_ws.cell(row=1, column=col)
        cell.value = "t1"

    # 2. Detect and mark zero values
    pink_fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')
    for row in merge_ws.iter_rows(min_row=2, max_row=merge_ws.max_row, max_col=max_col):
        for cell in row:
            if cell.value == 0 or cell.value == 0.0:
                cell.value = ""
                cell.fill = pink_fill
    # === END NEW CODE ===

    merge_path = os.path.join(target_folder, f"merge_{task_name}.xlsx")
    try:
        merge_wb.save(merge_path)
        print(f"Merged data saved to {merge_path}")
    except Exception as e:
        print(f"Error saving merged file: {str(e)}")

    print(f"Process completed for {task_name}. Files processed: {files_processed}, Sheets processed: {sheets_processed}")
    print(f"Total non-blank rows merged: {rows_merged}")
    return merge_path

# GUI Class for Combine Awards
class ExcelProcessorGUI:
    def __init__(self, root, file_path):
        self.root = root
        self.root.title("Excel Data Processor")
        self.root.geometry("800x600")
        self.file_path = file_path
        
        # Create main frame
        main_frame = ttk.Frame(root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Preview Section
        ttk.Label(main_frame, text="Preview of File:").grid(row=0, column=0, sticky=tk.W)
        self.preview_table = ttk.Treeview(main_frame, height=10)
        self.preview_table.grid(row=1, column=0, columnspan=3, pady=10)
        self.load_preview()
        
        # Sheet selection dropdown
        ttk.Label(main_frame, text="Sheet Name:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.sheet_combobox = ttk.Combobox(main_frame, width=47, state="readonly")
        self.sheet_combobox.grid(row=2, column=1, padx=5, pady=5)
        self.update_sheet_list()
        
        # Column indices
        ttk.Label(main_frame, text="Column Indices (specify by letter)").grid(row=3, column=0, columnspan=3, pady=10)
        
        labels = ["Class Name Column:", "Class Number Column:", "Grouping Criteria Column:", "Values to Combine Column:"]
        self.entries = []
        
        for i, label in enumerate(labels):
            ttk.Label(main_frame, text=label).grid(row=i+4, column=0, sticky=tk.W)
            entry = ttk.Entry(main_frame, width=10)
            entry.grid(row=i+4, column=1, sticky=tk.W, pady=5)
            self.entries.append(entry)
        
        # Process button
        ttk.Button(main_frame, text="Process Excel", command=self.process_excel).grid(row=8, column=0, columnspan=3, pady=20)

    def load_preview(self):
        # Load preview of the file
        try:
            df = pd.read_excel(self.file_path, nrows=10)  # Load first 10 rows for preview
            self.preview_table["columns"] = list(df.columns)
            self.preview_table["show"] = "headings"

            # Add headers with Excel-like column letters (A, B, C, etc.)
            for i, col in enumerate(df.columns):
                col_letter = get_column_letter(i + 1)
                self.preview_table.heading(col, text=f"{col_letter} - {col}")
                self.preview_table.column(col, width=100)
            for _, row in df.iterrows():
                self.preview_table.insert("", "end", values=list(row))
        except Exception as e:
            messagebox.showerror("Error", f"Error loading preview: {e}")

    def update_sheet_list(self):
        try:
            wb = load_workbook(self.file_path)
            sheet_names = wb.sheetnames
            self.sheet_combobox['values'] = sheet_names
            self.sheet_combobox.set(sheet_names[0])  # Set default value to first sheet
            wb.close()
        except Exception as e:
            messagebox.showerror("Error", f"Error loading Excel file:\n{str(e)}")

    def process_excel(self):
        try:
            # Get values from GUI
            sheet_name = self.sheet_combobox.get()
            col_letters = [entry.get().strip().upper() for entry in self.entries]  # Convert input to uppercase letters
            col_indices = [column_index_from_string(letter) - 1 for letter in col_letters]  # Convert to 0-based indices
            
            # Load workbook
            wb = load_workbook(self.file_path)
            ws = wb[sheet_name]
            
            # Create dictionary to store groups
            groups = {}
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            
            # First pass: identify groups and collect data
            rows_to_delete = set()
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):  # Skip header row
                try:
                    class_name = row[col_indices[0]].value.lower() if row[col_indices[0]].value else ""
                    class_num = row[col_indices[1]].value
                    group_criteria = row[col_indices[2]].value
                    
                    if not any([class_name, class_num, group_criteria]):
                        continue
                        
                    key = (class_name, class_num, group_criteria)
                    
                    if key not in groups:
                        groups[key] = {
                            'values': [row[col_indices[3]].value],
                            'first_row': row_idx,
                            'rows_to_combine': [row_idx]
                        }
                    else:
                        groups[key]['values'].append(row[col_indices[3]].value)
                        groups[key]['rows_to_combine'].append(row_idx)
                        rows_to_delete.add(row_idx)
                except IndexError:
                    messagebox.showerror("Error", f"Row {row_idx} has fewer columns than specified. Please check your column indices.")
                    return
                except Exception as e:
                    messagebox.showerror("Error", f"Error processing row {row_idx}: {str(e)}")
                    return
            
            # Second pass: combine values and delete rows
            for key, group_data in groups.items():
                if len(group_data['values']) > 1:
                    combined_value = ', '.join(v for v in group_data['values'] if v)
                    first_row = group_data['first_row']
                    
                    ws.cell(row=first_row, column=col_indices[3] + 1).value = combined_value
                    
                    for cell in ws[first_row]:
                        cell.fill = yellow_fill
            
            for row_idx in sorted(rows_to_delete, reverse=True):
                ws.delete_rows(row_idx)
            
            output_path = self.file_path.rsplit('.', 1)[0] + '_processed.xlsx'
            wb.save(output_path)
            
            messagebox.showinfo("Success", f"Processing complete!\nSaved as: {output_path}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()
    
    target_folder = filedialog.askdirectory(title="Select Target Folder")
    if not target_folder:
        print("No folder selected. Exiting.")
        exit()

    ole_process_range, ole_criteria_range = get_user_input("OLE")
    displine_process_range, displine_criteria_range = get_user_input("Displine")

    merge_ole_path = None
    if ole_process_range and ole_criteria_range:
        merge_ole_path = process_excel_files(target_folder, "OLE", ole_process_range, ole_criteria_range)

    if displine_process_range and displine_criteria_range:
        process_excel_files(target_folder, "Displine", displine_process_range, displine_criteria_range)

    if merge_ole_path:
        root = tk.Tk()
        app = ExcelProcessorGUI(root, merge_ole_path)
        root.mainloop()
