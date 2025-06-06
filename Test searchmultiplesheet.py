#!/usr/bin/env python
# coding: utf-8

# In[9]:


import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import threading
import win32com.client
import pythoncom
import time
import csv

def search_string_in_excel_files(search_strings, folder_path, result_callback, status_callback):
    start_time = time.time()
    results = []
    # Convert all search strings to lowercase and strip whitespace
    search_strings = [s.strip().lower() for s in search_strings.split(',')]
    
    # Update status to "Searching..."
    app.after(0, status_callback, "Searching...")

    for root, dirs, files in os.walk(folder_path):
        for filename in files:
            if filename.endswith((".xlsx", ".xlsm")) and not filename.startswith("~$"):
                file_path = os.path.join(root, filename)
                print(f"Attempting to search in file: {file_path}")
                try:
                    workbook = load_workbook(file_path, data_only=True, read_only=True)
                    
                    for sheet_name in workbook.sheetnames:
                        worksheet = workbook[sheet_name]
                        print(f"Searching in sheet: {sheet_name}")
                        
                        # Get values of H3, H4, H5,H17
                        h3_value = worksheet['H3'].value if worksheet['H3'].value else ''
                        h4_value = worksheet['H4'].value if worksheet['H4'].value else ''
                        h5_value = worksheet['H5'].value if worksheet['H5'].value else ''
                        h17_value = worksheet['H17'].value if worksheet['H17'].value else ''
                        
                        for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                            for col_idx, cell in enumerate(row, start=1):
                                if isinstance(cell, str):
                                    cell_lower = cell.lower()
                                    # Check if any of the search strings match
                                    if any(search_str in cell_lower for search_str in search_strings):
                                        print(f"Match found: {cell}")
                                        results.append((file_path, sheet_name, cell, row_idx, col_idx, h3_value, h4_value, h5_value, h17_value))
                    
                    workbook.close()
                except (PermissionError, InvalidFileException) as e:
                    print(f"Could not open file {file_path}. Error: {e}")
                    continue

    end_time = time.time()
    duration = round(end_time - start_time, 2)
    
    # Call the result_callback on the main thread using app.after
    app.after(0, result_callback, results)
    
    # Show completion message with search duration
    completion_message = f"Search completed in {duration} seconds"
    app.after(0, status_callback, completion_message)
    app.after(0, lambda: messagebox.showinfo("Search Complete", 
                                           f"Search has been completed!\n\n"
                                           f"Total matches found: {len(results)}\n"
                                           f"Time taken: {duration} seconds"))

def browse_folder():
    folder_path = filedialog.askdirectory()
    folder_entry.delete(0, tk.END)
    folder_entry.insert(0, folder_path)

def update_status(message):
    status_label.config(text=message)

def clear_logs():
    # Clear the console output (if any)
    print("\n" * 100)  # Simple way to clear console
    # Clear the results tree
    for i in tree.get_children():
        tree.delete(i)
    # Reset status and count
    status_label.config(text="Ready")
    result_count.set("Total matches found: 0")

def export_to_csv():
    if not tree.get_children():
        messagebox.showwarning("No Data", "There are no results to export!")
        return
    
    try:
        # Ask user for save location
        file_path = filedialog.asksaveasfilename(
            defaultextension='.csv',
            filetypes=[("CSV files", "*.csv")],
            title="Export Results as CSV"
        )
        
        if not file_path:  # If user cancels the save dialog
            return
        
        # Get current timestamp
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
        
        # Write to CSV
        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Write header with timestamp
            writer.writerow(["Export Date (UTC):", timestamp])
            writer.writerow([])  # Empty row for spacing
            
            # Write column headers
            writer.writerow(["Filename", "Sheet Name", "Cell Content", "Row", "Column", "H3", "H4", "H5"])
            
            # Write data
            for item in tree.get_children():
                writer.writerow(tree.item(item)['values'])
        
        messagebox.showinfo("Export Successful", f"Results have been exported to:\n{file_path}")
        
    except Exception as e:
        messagebox.showerror("Export Error", f"An error occurred while exporting:\n{str(e)}")

def display_results(results):
    for i in tree.get_children():
        tree.delete(i)
    
    for idx, (file_path, sheet_name, cell, row, col, h3, h4, h5) in enumerate(results):
        tree.insert("", "end", iid=idx, values=(os.path.basename(file_path), sheet_name, cell, row, col, h3, h4, h5))
    
    result_count.set(f"Total matches found: {len(results)}")

def search():
    search_string = search_entry.get()
    folder_path = folder_entry.get()
    
    if not search_string or not folder_path:
        messagebox.showerror("Error", "Both search string and folder path are required.")
        return
    
    # Clear previous results and logs
    clear_logs()
    
    # Start a new thread for the search
    search_thread = threading.Thread(target=search_string_in_excel_files, 
                                   args=(search_string, folder_path, display_results, update_status))
    search_thread.start()

def on_double_click(event):
    try:
        item = tree.selection()[0]
        file_path, sheet_name, _, row, col, _, _, _ = tree.item(item, "values")
        full_path = os.path.join(folder_entry.get(), file_path)
        
        # Initialize COM in this thread
        pythoncom.CoInitialize()
        
        try:
            # Create Excel Application object with explicit visibility setting
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = 1  # Force visibility to True
            
            # Add error checking for file existence
            if not os.path.exists(full_path):
                raise FileNotFoundError(f"File not found: {full_path}")
            
            # Open workbook with full path
            workbook = excel.Workbooks.Open(os.path.abspath(full_path))
            
            # Add error checking for sheet existence
            try:
                worksheet = workbook.Worksheets(sheet_name)
            except:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
            
            worksheet.Activate()
            cell = worksheet.Cells(int(row), int(col))
            cell.Select()
            
            messagebox.showinfo("File Opened", f"Opened {file_path}\nSheet: {sheet_name}\nCell: R{row}C{col}")
            
        except Exception as e:
            # Release Excel application in case of error
            try:
                excel.Quit()
            except:
                pass
            messagebox.showerror("Error", f"Could not open file or navigate to cell:\n{str(e)}")
            
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred:\n{str(e)}")
    
    finally:
        # Always uninitialize COM
        pythoncom.CoUninitialize()

# Set up the main application window
app = tk.Tk()
app.title("Excel Search Tool")

frame = tk.Frame(app)
frame.pack(pady=20)

tk.Label(frame, text="Search Strings (separate by comma):").grid(row=0, column=0, padx=5, pady=5)
search_entry = tk.Entry(frame, width=50)
search_entry.grid(row=0, column=1, padx=5, pady=5)

tk.Label(frame, text="Folder Path:").grid(row=1, column=0, padx=5, pady=5)
folder_entry = tk.Entry(frame, width=50)
folder_entry.grid(row=1, column=1, padx=5, pady=5)
tk.Button(frame, text="Browse", command=browse_folder).grid(row=1, column=2, padx=5, pady=5)

# Add a help label for search syntax
help_text = "Example: Enter 'GGG,hhh' to search for cells containing either 'GGG' or 'hhh'"
help_label = tk.Label(app, text=help_text, fg="gray")
help_label.pack(pady=5)

# Frame for buttons
button_frame = tk.Frame(app)
button_frame.pack(pady=10)

tk.Button(button_frame, text="Search", command=search).pack(side=tk.LEFT, padx=5)
tk.Button(button_frame, text="Export to CSV", command=export_to_csv).pack(side=tk.LEFT, padx=5)

result_count = tk.StringVar()
tk.Label(app, textvariable=result_count).pack()

# Add status label
status_label = tk.Label(app, text="Ready", font=("Arial", 10))
status_label.pack(pady=5)

columns = ("Filename", "Sheet Name", "Cell Content", "Row", "Column", "H3", "H4", "H5")
tree = ttk.Treeview(app, columns=columns, show='headings')
for col in columns:
    tree.heading(col, text=col)
    tree.column(col, width=100)

# Add vertical scrollbar
scrollbar = ttk.Scrollbar(app, orient="vertical", command=tree.yview)
tree.configure(yscrollcommand=scrollbar.set)

# Pack the treeview and scrollbar
tree.pack(side="left", fill="both", expand=True, pady=20)
scrollbar.pack(side="right", fill="y")

tree.bind("<Double-1>", on_double_click)

app.mainloop()

