import openpyxl
import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog
from tkinter import ttk

# Function to handle sheet selection and copying
def copy_selected_sheets():
    selected_sheets = [sheet for sheet, var in zip(sheet_names, check_vars) if var.get()]
    
    if not selected_sheets:
        messagebox.showwarning("No Selection", "Please select at least one sheet to copy.")
        return
    
    # Create a new workbook for output
    output_wb = openpyxl.Workbook()
    default_sheet = output_wb.active
    output_wb.remove(default_sheet)  # Remove the default sheet in the new workbook
    
    # Copy the selected sheets to the new workbook
    for sheet_name in selected_sheets:
        input_sheet = input_wb[sheet_name]
        new_sheet = output_wb.create_sheet(title=sheet_name)
        
        # Copy the sheet contents
        for row in input_sheet.iter_rows(values_only=True):
            new_sheet.append(row)
    
    # Save the new workbook
    output_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", title="Save Output Workbook",
                                                    filetypes=[("Excel files", "*.xlsx")])
    if output_file_path:
        output_wb.save(output_file_path)
        messagebox.showinfo("Success", f"Selected sheets copied successfully to {output_file_path}")

# Function to create the GUI with a scrollable list of checkboxes
def create_gui(sheet_names):
    root = tk.Tk()
    root.title("Select Sheets to Copy")
    
    # Create a canvas to enable scrolling
    canvas = tk.Canvas(root)
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    
    # Create a vertical scrollbar linked to the canvas
    scrollbar = ttk.Scrollbar(root, orient=tk.VERTICAL, command=canvas.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    # Configure the canvas to work with the scrollbar
    canvas.configure(yscrollcommand=scrollbar.set)
    canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    
    # Create a frame inside the canvas to hold the checkboxes
    frame = tk.Frame(canvas)
    canvas.create_window((0, 0), window=frame, anchor="nw")
    
    # Label
    tk.Label(frame, text="Select sheets to copy:").pack(anchor="w")
    
    # Create a list of checkboxes for each sheet
    global check_vars
    check_vars = []
    
    for sheet_name in sheet_names:
        var = tk.BooleanVar()
        check_vars.append(var)
        tk.Checkbutton(frame, text=sheet_name, variable=var).pack(anchor="w")
    
    # Create a button to trigger the sheet copying
    tk.Button(root, text="Copy Selected Sheets", command=copy_selected_sheets).pack(pady=10)
    
    # Set the canvas scroll region and configure the scrollbar
    frame.update_idletasks()
    canvas.config(scrollregion=canvas.bbox("all"))
    
    root.mainloop()

# Load the input workbook
input_file_path = filedialog.askopenfilename(title="Select Excel Workbook", filetypes=[("Excel files", "*.xlsx")])
if not input_file_path:
    messagebox.showerror("Error", "No file selected. Exiting.")
    exit()

input_wb = openpyxl.load_workbook(input_file_path)
sheet_names = input_wb.sheetnames

# Launch the GUI
create_gui(sheet_names)