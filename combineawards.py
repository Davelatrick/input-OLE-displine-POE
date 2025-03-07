import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from openpyxl.styles import PatternFill

class ExcelProcessorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Data Processor")
        self.root.geometry("600x500")
        
        # Create main frame
        main_frame = ttk.Frame(root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # File selection
        ttk.Label(main_frame, text="Excel File:").grid(row=0, column=0, sticky=tk.W)
        self.file_path = tk.StringVar()
        self.file_entry = ttk.Entry(main_frame, textvariable=self.file_path, width=50)
        self.file_entry.grid(row=0, column=1, padx=5)
        ttk.Button(main_frame, text="Browse", command=self.browse_file).grid(row=0, column=2)
        
        # Sheet selection dropdown
        ttk.Label(main_frame, text="Sheet Name:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.sheet_combobox = ttk.Combobox(main_frame, width=47, state="readonly")
        self.sheet_combobox.grid(row=1, column=1, padx=5, pady=5)
        
        # Column indices
        ttk.Label(main_frame, text="Column Indices (starting from 1)").grid(row=2, column=0, columnspan=3, pady=10)
        
        labels = ["Class Name Column:", "Class Number Column:", "Grouping Criteria Column:", "Values to Combine Column:"]
        self.entries = []
        
        for i, label in enumerate(labels):
            ttk.Label(main_frame, text=label).grid(row=i+3, column=0, sticky=tk.W)
            entry = ttk.Entry(main_frame, width=10)
            entry.grid(row=i+3, column=1, sticky=tk.W, pady=5)
            self.entries.append(entry)
        
        # Process button
        ttk.Button(main_frame, text="Process Excel", command=self.process_excel).grid(row=7, column=0, columnspan=3, pady=20)

    def browse_file(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
        if filename:
            self.file_path.set(filename)
            self.update_sheet_list()

    def update_sheet_list(self):
        try:
            wb = openpyxl.load_workbook(self.file_path.get())
            sheet_names = wb.sheetnames
            self.sheet_combobox['values'] = sheet_names
            self.sheet_combobox.set(sheet_names[0])  # Set default value to first sheet
            wb.close()
        except Exception as e:
            messagebox.showerror("Error", f"Error loading Excel file:\n{str(e)}")

    def safe_get_value(self, cell):
        """Safely get cell value, converting to string and handling None"""
        if cell.value is None:
            return ""
        return str(cell.value).strip()

    def process_excel(self):
        try:
            # Get values from GUI
            file_path = self.file_path.get()
            sheet_name = self.sheet_combobox.get()
            col_indices = [int(entry.get()) - 1 for entry in self.entries]  # Convert to 0-based indices
            
            # Load workbook
            wb = openpyxl.load_workbook(file_path)
            ws = wb[sheet_name]
            
            # Create dictionary to store groups
            groups = {}
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            
            # First pass: identify groups and collect data
            rows_to_delete = set()
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):  # Skip header row
                # Safely get values with error handling
                try:
                    class_name = self.safe_get_value(row[col_indices[0]]).lower()
                    class_num = self.safe_get_value(row[col_indices[1]])
                    group_criteria = self.safe_get_value(row[col_indices[2]])
                    
                    # Skip completely empty rows
                    if not any([class_name, class_num, group_criteria]):
                        continue
                        
                    key = (class_name, class_num, group_criteria)
                    
                    if key not in groups:
                        groups[key] = {
                            'values': [self.safe_get_value(row[col_indices[3]])],
                            'first_row': row_idx,
                            'rows_to_combine': [row_idx]
                        }
                    else:
                        groups[key]['values'].append(self.safe_get_value(row[col_indices[3]]))
                        groups[key]['rows_to_combine'].append(row_idx)
                        rows_to_delete.add(row_idx)
                except IndexError:
                    messagebox.showerror("Error", f"Row {row_idx} has fewer columns than specified. Please check your column indices.")
                    return
                except Exception as e:
                    messagebox.showerror("Error", f"Error processing row {row_idx}: {str(e)}")
                    return
            
            # Second pass: combine values and delete rows
            for key, group_data in groups.items():
                if len(group_data['values']) > 1:
                    # Combine non-empty values
                    combined_value = ', '.join(v for v in group_data['values'] if v)
                    first_row = group_data['first_row']
                    
                    # Update the combined value
                    ws.cell(row=first_row, column=col_indices[3] + 1).value = combined_value
                    
                    # Color the entire row yellow
                    for cell in ws[first_row]:
                        cell.fill = yellow_fill
            
            # Delete rows from bottom to top to avoid shifting issues
            for row_idx in sorted(rows_to_delete, reverse=True):
                ws.delete_rows(row_idx)
            
            # Save the workbook
            output_path = file_path.rsplit('.', 1)[0] + '_processed.xlsx'
            wb.save(output_path)
            
            messagebox.showinfo("Success", f"Processing complete!\nSaved as: {output_path}")
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelProcessorGUI(root)
    root.mainloop()