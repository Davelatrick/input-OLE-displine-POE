import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill
import pandas as pd
import logging

# Set up logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('script_log.txt'),
        logging.StreamHandler()
    ]
)

# Function to parse range strings
def parse_range(range_str):
    start, end = range_str.split(':')
    start_col, start_row = start[:2], int(start[2:])
    end_col, end_row = end[:2], int(end[2:])
    return start_col, start_row, end_col, end_row

# Function to process Excel files and generate merged files
def process_excel_files(target_folder, task_name, process_range, criteria_range):
    logging.info(f"Starting the process for {task_name}...")
    print(f"Starting the process for {task_name}...")
    print(f"Process range: {process_range}")
    print(f"Criteria range: {criteria_range}")

    proc_start_col, proc_start_row, proc_end_col, proc_end_row = parse_range(process_range)
    crit_start_col, crit_start_row, crit_end_col, crit_end_row = parse_range(criteria_range)

    merge_wb = Workbook()
    merge_ws = merge_wb.active
    merge_ws.title = f"Merged Data {task_name}"

    files_processed = 0
    sheets_processed = 0
    rows_merged = 0

    for filename in os.listdir(target_folder):
        if filename.endswith('.xlsx') and not filename.startswith('~$'):
            file_path = os.path.join(target_folder, filename)
            logging.info(f"Processing file: {filename}")
            print(f"Processing file: {filename}")

            try:
                wb = load_workbook(file_path, data_only=True)

                for sheet_name in wb.sheetnames:
                    if sheet_name not in ["index", "list", "setting", "TEMPLATE", "STUDENTINFO"]:
                        logging.info(f"  Processing sheet: {sheet_name}")
                        print(f"  Processing sheet: {sheet_name}")
                        ws = wb[sheet_name]

                        last_row = crit_start_row - 1
                        for row in range(crit_start_row, crit_end_row + 1):
                            if ws[f'{crit_start_col}{row}'].value is not None and ws[f'{crit_start_col}{row}'].value != "":
                                last_row = row

                        if last_row >= crit_start_row:
                            logging.info(f"    Copying range {proc_start_col}{proc_start_row}:{proc_end_col}{last_row}")
                            print(f"    Copying range {proc_start_col}{proc_start_row}:{proc_end_col}{last_row}")

                            # Copy non-blank rows to merge sheet
                            for row in range(proc_start_row, last_row + 1):
                                if any(ws.cell(row=row, column=col).value not in (None, "")
                                       for col in range(column_index_from_string(proc_start_col), column_index_from_string(proc_end_col) + 1)):
                                    rows_merged += 1
                                    for col in range(column_index_from_string(proc_start_col), column_index_from_string(proc_end_col) + 1):
                                        source_value = ws.cell(row=row, column=col).value
                                        merge_ws.cell(row=rows_merged, column=col - column_index_from_string(proc_start_col) + 1, value=source_value)

                                    # Add sheet name in column S
                                    merge_ws.cell(row=rows_merged, column=19, value=sheet_name)

                            sheets_processed += 1

                wb.close()
                files_processed += 1

            except Exception as e:
                logging.error(f"Error processing file {filename}: {str(e)}")
                print(f"Error processing file {filename}: {str(e)}")

    # Concatenate values from columns P and S and replace values in column P
    for row in range(1, merge_ws.max_row + 1):
        p_value = merge_ws.cell(row=row, column=16).value
        s_value = merge_ws.cell(row=row, column=19).value
        if p_value and s_value:
            concat_value = f"{p_value}-{s_value}"
            merge_ws.cell(row=row, column=16, value=concat_value)

    # Additional concatenation for "Displine" task: concatenate I and S into I
    if task_name == "Displine":
        for row in range(1, merge_ws.max_row + 1):
            i_value = merge_ws.cell(row=row, column=9).value  # Column I
            s_value = merge_ws.cell(row=row, column=19).value  # Column S
            if i_value and s_value:
                concat_value = f"{i_value}-{s_value}"
                merge_ws.cell(row=row, column=9, value=concat_value)

    # Insert header row with sequential labels T1, T2, T3, etc.
    merge_ws.insert_rows(1)  # Insert new row at the top
    max_column = merge_ws.max_column  # Get rightmost column
    for col in range(1, max_column + 1):  # Label columns T1, T2, etc.
        merge_ws.cell(row=1, column=col).value = f"T{col}"

    # Replace 0 values with blank and apply pink fill
    pink_fill = PatternFill(start_color='FFC1CC', end_color='FFC1CC', fill_type='solid')
    for row in merge_ws.iter_rows(min_row=2, max_row=merge_ws.max_row, max_col=merge_ws.max_column):  # Skip header row
        for cell in row:
            if cell.value == 0:  # Check for exact 0 (integer or float)
                cell.value = None  # Replace with blank
                cell.fill = pink_fill  # Apply pink fill

    merge_path = os.path.join(target_folder, f"merge_{task_name}.xlsx")
    try:
        merge_wb.save(merge_path)
        logging.info(f"Merged data saved to {merge_path}")
        print(f"Merged data saved to {merge_path}")
    except Exception as e:
        logging.error(f"Error saving merged file: {str(e)}")
        print(f"Error saving merged file: {str(e)}")

    logging.info(f"Process completed for {task_name}. Files processed: {files_processed}, Sheets processed: {sheets_processed}")
    logging.info(f"Total non-blank rows merged: {rows_merged}")
    print(f"Process completed for {task_name}. Files processed: {files_processed}, Sheets processed: {sheets_processed}")
    print(f"Total non-blank rows merged: {rows_merged}")
    return merge_path

# GUI Class for Combine Awards
class ExcelProcessorGUI:
    def __init__(self, root, file_path, ole_process_range, ole_criteria_range, displine_process_range, displine_criteria_range):
        self.root = root
        self.root.title("Excel Data Processor")
        self.root.geometry("800x600")
        self.file_path = file_path
        self.ole_process_range = ole_process_range
        self.ole_criteria_range = ole_criteria_range
        self.displine_process_range = displine_process_range
        self.displine_criteria_range = displine_criteria_range
        
        # Create main frame
        main_frame = ttk.Frame(root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Preview Section
        ttk.Label(main_frame, text="Preview of File:").grid(row=0, column=0, sticky=tk.W)
        self.preview_table = ttk.Treeview(main_frame, height=10)
        self.preview_table.grid(row=1, column=0, columnspan=3, pady=10)
        self.load_preview()
        
        # Sheet selection dropdown
        ttk.Label(main_frame, text="Sheet Name:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.sheet_combobox = ttk.Combobox(main_frame, width=47, state="readonly")
        self.sheet_combobox.grid(row=2, column=1, padx=5, pady=5)
        self.update_sheet_list()
        
        # Range inputs for OLE and Displine
        ttk.Label(main_frame, text="Range Inputs (edit if needed)").grid(row=3, column=0, columnspan=3, pady=10)
        
        range_labels = [
            "OLE Process Range (e.g., CV26:DM205):",
            "OLE Criteria Range (e.g., CV26:CV205):",
            "Displine Process Range (e.g., BB26:BJ205):",
            "Displine Criteria Range (e.g., BB26:BB205):"
        ]
        self.range_entries = []
        default_ranges = [self.ole_process_range, self.ole_criteria_range, self.displine_process_range, self.displine_criteria_range]
        
        for i, label in enumerate(range_labels):
            ttk.Label(main_frame, text=label).grid(row=i+4, column=0, sticky=tk.W)
            entry = ttk.Entry(main_frame, width=20)
            entry.insert(0, default_ranges[i])  # Pre-populate with default values
            entry.grid(row=i+4, column=1, sticky=tk.W, pady=5)
            self.range_entries.append(entry)
        
        # Column indices
        ttk.Label(main_frame, text="Column Indices (specify by letter)").grid(row=8, column=0, columnspan=3, pady=10)
        
        labels = ["Class Name Column:", "Class Number Column:", "Grouping Criteria Column:", "Values to Combine Column:"]
        self.entries = []
        
        for i, label in enumerate(labels):
            ttk.Label(main_frame, text=label).grid(row=i+9, column=0, sticky=tk.W)
            entry = ttk.Entry(main_frame, width=10)
            entry.grid(row=i+9, column=1, sticky=tk.W, pady=5)
            self.entries.append(entry)
        
        # Process button
        ttk.Button(main_frame, text="Process Excel", command=self.process_excel).grid(row=13, column=0, columnspan=3, pady=20)

    def load_preview(self):
        try:
            df = pd.read_excel(self.file_path, nrows=10)  # Load first 10 rows for preview
            self.preview_table["columns"] = list(df.columns)
            self.preview_table["show"] = "headings"

            for i, col in enumerate(df.columns):
                col_letter = get_column_letter(i + 1)
                self.preview_table.heading(col, text=f"{col_letter} - {col}")
                self.preview_table.column(col, width=100)
            for _, row in df.iterrows():
                self.preview_table.insert("", "end", values=list(row))
        except Exception as e:
            logging.error(f"Error loading preview: {e}")
            messagebox.showerror("Error", f"Error loading preview: {e}")

    def update_sheet_list(self):
        try:
            wb = load_workbook(self.file_path)
            sheet_names = wb.sheetnames
            self.sheet_combobox['values'] = sheet_names
            self.sheet_combobox.set(sheet_names[0])  # Set default value to first sheet
            wb.close()
        except Exception as e:
            logging.error(f"Error loading Excel file: {str(e)}")
            messagebox.showerror("Error", f"Error loading Excel file:\n{str(e)}")

    def process_excel(self):
        try:
            sheet_name = self.sheet_combobox.get()
            col_letters = [entry.get().strip().upper() for entry in self.entries]  # Convert input to uppercase letters
            col_indices = [column_index_from_string(letter) - 1 for letter in col_letters]  # Convert to 0-based indices
            
            wb = load_workbook(self.file_path)
            ws = wb[sheet_name]
            
            groups = {}
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            
            rows_to_delete = set()
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):  # Skip header row
                try:
                    class_name = row[col_indices[0]].value.lower() if row[col_indices[0]].value else ""
                    class_num = row[col_indices[1]].value
                    group_criteria = row[col_indices[2]].value
                    
                    if not any([class_name, class_num, group_criteria]):
                        continue
                        
                    key = (class_name, class_num, group_criteria)
                    
                    if key not in groups:
                        groups[key] = {
                            'values': [row[col_indices[3]].value],
                            'first_row': row_idx,
                            'rows_to_combine': [row_idx]
                        }
                    else:
                        groups[key]['values'].append(row[col_indices[3]].value)
                        groups[key]['rows_to_combine'].append(row_idx)
                        rows_to_delete.add(row_idx)
                except IndexError:
                    logging.error(f"Row {row_idx} has fewer columns than specified. Please check your column indices.")
                    messagebox.showerror("Error", f"Row {row_idx} has fewer columns than specified. Please check your column indices.")
                    return
                except Exception as e:
                    logging.error(f"Error processing row {row_idx}: {str(e)}")
                    messagebox.showerror("Error", f"Error processing row {row_idx}: {str(e)}")
                    return
            
            for key, group_data in groups.items():
                if len(group_data['values']) > 1:
                    combined_value = ', '.join(v for v in group_data['values'] if v)
                    first_row = group_data['first_row']
                    
                    ws.cell(row=first_row, column=col_indices[3] + 1).value = combined_value
                    
                    for cell in ws[first_row]:
                        cell.fill = yellow_fill
            
            for row_idx in sorted(rows_to_delete, reverse=True):
                ws.delete_rows(row_idx)
            
            output_path = self.file_path.rsplit('.', 1)[0] + '_processed.xlsx'
            wb.save(output_path)
            
            logging.info(f"Processing complete! Saved as: {output_path}")
            messagebox.showinfo("Success", f"Processing complete!\nSaved as: {output_path}")
        except Exception as e:
            logging.error(f"Error in process_excel: {str(e)}")
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")

if __name__ == "__main__":
    try:
        logging.info("Starting script...")
        root = tk.Tk()
        root.withdraw()
        logging.info("Tkinter root initialized")

        target_folder = filedialog.askdirectory(title="Select Target Folder")
        if not target_folder:
            logging.error("No folder selected. Exiting.")
            print("No folder selected. Exiting.")
            exit()

        logging.info(f"Selected folder: {target_folder}")

        # Define default ranges
        ole_process_range = "CV26:DM205"
        ole_criteria_range = "CV26:CV205"
        displine_process_range = "BB26:BJ205"
        displine_criteria_range = "BB26:BB205"

        merge_ole_path = None
        if ole_process_range and ole_criteria_range:
            logging.info("Processing OLE...")
            merge_ole_path = process_excel_files(target_folder, "OLE", ole_process_range, ole_criteria_range)
        else:
            logging.warning("OLE ranges invalid or skipped")

        if displine_process_range and displine_criteria_range:
            logging.info("Processing Displine...")
            process_excel_files(target_folder, "Displine", displine_process_range, displine_criteria_range)
        else:
            logging.warning("Displine ranges invalid or skipped")

        if merge_ole_path:
            logging.info("Launching GUI with merge_ole_path")
            root = tk.Tk()
            app = ExcelProcessorGUI(root, merge_ole_path, ole_process_range, ole_criteria_range, displine_process_range, displine_criteria_range)
            root.mainloop()
        else:
            logging.error("No merge_ole_path generated. GUI not launched.")

    except Exception as e:
        logging.exception(f"Script failed: {str(e)}")
        print(f"Error: {str(e)}")
        messagebox.showerror("Error", f"Script failed: {str(e)}")