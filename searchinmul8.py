import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import threading
import win32com.client
import pythoncom

def search_string_in_excel_files(search_string, folder_path, result_callback):
    results = []
    search_string = search_string.lower()  # Convert search string to lowercase

    for root, dirs, files in os.walk(folder_path):
        for filename in files:
            if filename.endswith((".xlsx", ".xlsm")) and not filename.startswith("~$"):
                file_path = os.path.join(root, filename)
                print(f"Attempting to search in file: {file_path}")
                try:
                    workbook = load_workbook(file_path, data_only=True, read_only=True)
                    
                    for sheet_name in workbook.sheetnames:
                        worksheet = workbook[sheet_name]
                        print(f"Searching in sheet: {sheet_name}")
                        
                        # Get values of H3, H4, H5
                        h3_value = worksheet['H3'].value if worksheet['H3'].value else ''
                        h4_value = worksheet['H4'].value if worksheet['H4'].value else ''
                        h5_value = worksheet['H5'].value if worksheet['H5'].value else ''
                        
                        for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                            for col_idx, cell in enumerate(row, start=1):
                                if isinstance(cell, str) and search_string in cell.lower():  # Case-insensitive search
                                    print(f"Match found: {cell}")
                                    results.append((file_path, sheet_name, cell, row_idx, col_idx, h3_value, h4_value, h5_value))
                    
                    workbook.close()
                except (PermissionError, InvalidFileException) as e:
                    print(f"Could not open file {file_path}. Error: {e}")
                    continue

    # Call the result_callback on the main thread using app.after
    app.after(0, result_callback, results)

def browse_folder():
    folder_path = filedialog.askdirectory()
    folder_entry.delete(0, tk.END)
    folder_entry.insert(0, folder_path)

def display_results(results):
    for i in tree.get_children():
        tree.delete(i)
    
    for idx, (file_path, sheet_name, cell, row, col, h3, h4, h5) in enumerate(results):
        tree.insert("", "end", iid=idx, values=(os.path.basename(file_path), sheet_name, cell, row, col, h3, h4, h5))
    
    result_count.set(f"Total matches found: {len(results)}")

def search():
    search_string = search_entry.get()
    folder_path = folder_entry.get()
    
    if not search_string or not folder_path:
        messagebox.showerror("Error", "Both search string and folder path are required.")
        return
    
    # Start a new thread for the search
    search_thread = threading.Thread(target=search_string_in_excel_files, args=(search_string, folder_path, display_results))
    search_thread.start()

def on_double_click(event):
    item = tree.selection()[0]
    file_path, sheet_name, _, row, col, _, _, _ = tree.item(item, "values")
    full_path = os.path.join(folder_entry.get(), file_path)
    
    try:
        pythoncom.CoInitialize()
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = True
        workbook = excel.Workbooks.Open(os.path.abspath(full_path))
        worksheet = workbook.Worksheets(sheet_name)
        worksheet.Activate()
        cell = worksheet.Cells(int(row), int(col))
        cell.Select()
        messagebox.showinfo("File Opened", f"Opened {file_path}\nSheet: {sheet_name}\nCell: R{row}C{col}")
    except Exception as e:
        messagebox.showerror("Error", f"Could not open file or navigate to cell: {e}")
    finally:
        pythoncom.CoUninitialize()

# Set up the main application window
app = tk.Tk()
app.title("Excel Search Tool")

frame = tk.Frame(app)
frame.pack(pady=20)

tk.Label(frame, text="Search String:").grid(row=0, column=0, padx=5, pady=5)
search_entry = tk.Entry(frame, width=50)
search_entry.grid(row=0, column=1, padx=5, pady=5)

tk.Label(frame, text="Folder Path:").grid(row=1, column=0, padx=5, pady=5)
folder_entry = tk.Entry(frame, width=50)
folder_entry.grid(row=1, column=1, padx=5, pady=5)
tk.Button(frame, text="Browse", command=browse_folder).grid(row=1, column=2, padx=5, pady=5)

tk.Button(app, text="Search", command=search).pack(pady=10)

result_count = tk.StringVar()
tk.Label(app, textvariable=result_count).pack()

columns = ("Filename", "Sheet Name", "Cell Content", "Row", "Column", "H3", "H4", "H5")
tree = ttk.Treeview(app, columns=columns, show='headings')
for col in columns:
    tree.heading(col, text=col)
    tree.column(col, width=100)

# Add vertical scrollbar
scrollbar = ttk.Scrollbar(app, orient="vertical", command=tree.yview)
tree.configure(yscrollcommand=scrollbar.set)

# Pack the treeview and scrollbar
tree.pack(side="left", fill="both", expand=True, pady=20)
scrollbar.pack(side="right", fill="y")

tree.bind("<Double-1>", on_double_click)

app.mainloop()